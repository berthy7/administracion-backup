from servidor.common.managers import SuperManager
from servidor.sistema.usuarios.bitacora.manager import BitacoraManager
from servidor.sistema.usuarios.usuario.manager import UsuarioManager
from servidor.sistema.usuarios.bitacora.model import Bitacora
from servidor.sistema.almacen.descuento.model import Descuento
from servidor.sistema.recursos_humanos.motivo.manager import MotivoManager
from servidor.sistema.recursos_humanos.personal.model import Personal

from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from dateutil.relativedelta import relativedelta

import pytz
from sqlalchemy.sql import func
from sqlalchemy import func, desc, asc,or_,and_
from openpyxl import Workbook
from openpyxl.styles import Border , Alignment, Font, PatternFill,Side
from openpyxl.worksheet.table import Table, TableStyleInfo

global fecha_zona
fecha_zona = datetime.now(pytz.timezone('America/La_Paz'))


class DescuentoManager(SuperManager):

    def __init__(self, db):
        super().__init__(Descuento, db)


    def listar_x_personal(self,fkpersonal):
        return self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento").filter(self.entity.fkpersonal == fkpersonal).all()

    def listar_habilitados(self):
        return self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento").all()

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento").all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento"))

    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento")
        return items

    def all_data(self, idu):
        privilegios = UsuarioManager(self.db).get_privileges(idu, '/descuento')
        datos = self.db.query(self.entity).filter(self.entity.enabled).filter(self.entity.tipo == "Descuento").all()
        list = []

        for item in datos:
            is_active = item.estado
            disable = 'disabled' if 'descuento_update' not in privilegios else ''
            estado = 'Activo' if is_active else 'Inactivo'
            check = 'checked' if is_active else ''
            delete = 'descuento_delete' in privilegios

            diccionario = item.get_dict()
            diccionario['estado'] = estado
            diccionario['check'] = check
            diccionario['disable'] = disable
            diccionario['delete'] = delete
            diccionario['fecha'] = item.fecha.strftime('%d/%m/%Y %H:%M')
            diccionario['personal'] = item.personal.fullname
            diccionario['motivo'] = item.motivo.nombre
            list.append(diccionario)

        return list

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        objeto.fecha = fecha
        objeto.tipo = "Descuento"

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registr?? descuento.", fecha=fecha, tabla="almacen_descuento", identificador=a.id)
        super().insert(b)
        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modific?? descuento.", fecha=fecha, tabla="almacen_descuento", identificador=a.id)
        super().insert(b)
        return a

    def state(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        mensaje = "Habilit?? descuento" if estado else "Deshabilit?? descuento"
        x.estado = estado

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion=mensaje, fecha=fecha, tabla="almacen_descuento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

    def delete(self, id, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = False

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Elimin?? descuento", fecha=fecha, tabla="almacen_descuento", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()

    def descuento_excel(self, diccionary):
        fecha = datetime.now()
        cname = "Descuento" + fecha.strftime('%Y-%m-%d') + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de Desccuento'

        indice = 0

        # --------------------------------------------------------------------
        indice = indice + 1
        ws['A' + str(indice)] = 'PERSONAL'
        indice = indice + 2

        ws['A' + str(indice)] = 'N??'
        ws['B' + str(indice)] = 'CI'
        ws['C' + str(indice)] = 'FECHA INGRESO'
        ws['D' + str(indice)] = 'FECHA DE NACIMIENTO'
        ws['E' + str(indice)] = 'CARGO'
        ws['F' + str(indice)] = 'TELEFONO'
        ws['G' + str(indice)] = 'NOMBRE COMPLETO'

        ws['A' + str(indice)].font = Font(bold=True)
        ws['B' + str(indice)].font = Font(bold=True)
        ws['C' + str(indice)].font = Font(bold=True)
        ws['D' + str(indice)].font = Font(bold=True)
        ws['E' + str(indice)].font = Font(bold=True)
        ws['F' + str(indice)].font = Font(bold=True)
        ws['G' + str(indice)].font = Font(bold=True)


        array_indices = ['H', 'I','J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA',
                         'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP',
                         'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

        fechainicio = datetime.strptime(diccionary['fecha-inicio-reporte'], '%d/%m/%Y')
        fechafin = datetime.strptime(diccionary['fecha-fin-reporte'], '%d/%m/%Y')
        cont_fecha = 0
        motivos = MotivoManager(self.db).listar_x_descuento()

        list_dict_tipo_asistencia = []

        for tipo_asist in motivos:
            list_dict_tipo_asistencia.append(dict(id_asistencia=tipo_asist.id,tipo_asistencia=tipo_asist.nombre, cantidad=0))

            ws[array_indices[cont_fecha] + str(indice)] = tipo_asist.nombre
            ws[array_indices[cont_fecha] + str(indice)].font = Font(bold=True)


            cont_fecha = cont_fecha + 1

        ws[array_indices[cont_fecha] + str(indice)] = 'TOTAL Bs.'
        ws[array_indices[cont_fecha] + str(indice)].font = Font(bold=True)

        if len(diccionary['personal']) == 0:
            personal = self.db.query(Personal).filter(Personal.estado).filter(Personal.enabled).all()
        else:
            personal = self.db.query(Personal).filter(Personal.estado).filter(Personal.enabled).filter(
                Personal.id.in_(diccionary['personal'])).all()

        for per in personal:

            indice = indice + 1
            ws['A' + str(indice)] = str(per.id)
            ws['B' + str(indice)] = str(per.ci)
            ws['C' + str(indice)] = per.fechar.strftime('%d/%m/%Y')
            ws['D' + str(indice)] = per.fechanacimiento.strftime('%d/%m/%Y')
            ws['E' + str(indice)] = str(per.cargo.nombre)
            ws['F' + str(indice)] = str(per.telefono)
            ws['G' + str(indice)] = str(per.fullname)

            cont_fecha = 0
            total_sum_descuento = 0

            for tip_asistencia in list_dict_tipo_asistencia:
                descuento = self.db.query(func.sum(Descuento.monto)).filter(
                    tip_asistencia['id_asistencia'] == Descuento.fkmotivo).filter(Descuento.tipo == "Descuento").filter(
                    Descuento.fkpersonal == per.id)

                sum_descuento = descuento[0][0]
                if sum_descuento:
                    total_sum_descuento = total_sum_descuento + sum_descuento


                ws[array_indices[cont_fecha] + str(indice)] = tip_asistencia['cantidad']
                cont_fecha = cont_fecha + 1

            ws[array_indices[cont_fecha] + str(indice)] = total_sum_descuento

        for col in ws.columns:

            column = col[0].column_letter
            ws.column_dimensions[column].height = 1

            if column == "A":
                ws.column_dimensions[column].width = 5
            elif column == "B":
                ws.column_dimensions[column].width = 14
            elif column == "C":
                ws.column_dimensions[column].width = 16
            elif column == "D":
                ws.column_dimensions[column].width = 16
            elif column == "E":
                ws.column_dimensions[column].width = 22
            elif column == "F":
                ws.column_dimensions[column].width = 12
            elif column == "G":
                ws.column_dimensions[column].width = 30
            else:
                ws.column_dimensions[column].width = 11

        wb.save("servidor/common/resources/downloads/" + cname)

        return cname



