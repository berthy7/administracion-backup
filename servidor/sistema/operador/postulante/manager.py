from servidor.common.managers import SuperManager
from servidor.sistema.usuarios.bitacora.manager import BitacoraManager
from servidor.sistema.usuarios.usuario.manager import UsuarioManager
from servidor.sistema.usuarios.bitacora.model import Bitacora
from servidor.sistema.recursos_humanos.personal.model import Personal
from servidor.sistema.recursos_humanos.cargo.model import Cargo
from servidor.sistema.recursos_humanos.cargo.manager import CargoManager
from sqlalchemy.exc import IntegrityError
from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, desc, asc


import pytz

global fecha_zona
fecha_zona = datetime.now(pytz.timezone('America/La_Paz'))


class PostulanteManager(SuperManager):

    def __init__(self, db):
        super().__init__(Personal, db)

    def importar_excel(self, cname, user, ip):
        try:
            wb = load_workbook(filename="servidor/common/resources/uploads/" + cname)
            ws = wb.active
            colnames = ['NRO', 'CI', 'EXP', 'FECHA_INGRESO','FECHA_NACIMIENTO','CARGO','TELEFONO','APELLIDO_PATERNO','APELLIDO_MATERNO','NOMBRES']
            indices = {cell[0].value: n - 1 for n, cell in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1) if
                       cell[0].value in colnames}
            if len(indices) == len(colnames):
                for row in ws.iter_rows(min_row=2):

                    ci = str(row[indices['CI']].value) + " " + str(row[indices['EXP']].value)

                    query = self.db.query(Personal).filter(Personal.ci == ci).first()

                    if not query:

                        dicc_foto = dict()
                        list_doc = list()
                        list_doc.append(dicc_foto)

                        cargoNombre = row[indices['CARGO']].value



                        persona = Personal(nombre=str(row[indices['NOMBRES']].value),
                                          apellidop=str(row[indices['APELLIDO_PATERNO']].value),
                                          apellidom=str(row[indices['APELLIDO_MATERNO']].value),
                                          telefono=str(row[indices['TELEFONO']].value),
                                          ci=str(row[indices['CI']].value) + " " + row[indices['EXP']].value,
                                          fechanacimiento=row[indices['FECHA_NACIMIENTO']].value,
                                           fechar=row[indices['FECHA_INGRESO']].value,
                                           administrativos= [],
                                          familiares= [],
                                           experiencias=[],
                                            estudios=[],
                                            complementos= [],
                                           documentos=list_doc
                                          )


                        if CargoManager(self.db).obtener_x_nombre(cargoNombre) is None:
                            cargo = Cargo(nombre=cargoNombre)
                            persona.cargo = cargo
                        else:
                            persona.cargo = CargoManager(self.db).obtener_x_nombre(cargoNombre)

                        self.db.add(persona)


                self.db.commit()
                return {'message': 'Importado Todos Correctamente.', 'success': True}
            else:
                return {'message': 'Columnas Faltantes', 'success': False}
        except IntegrityError as e:
            self.db.rollback()
            if 'UNIQUE constraint' in str(e):
                return {'message': 'duplicado', 'success': False}
            if 'UNIQUE constraint failed' in str(e):
                return {'message': 'codigo duplicado', 'success': False}
            return {'message': str(e), 'success': False}

    def listar_habilitados(self):
        return self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).order_by(asc(self.entity.apellidop)).all()

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled).order_by(asc(self.entity.apellidop)).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.enabled).order_by(asc(self.entity.apellidop)))

    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).order_by(asc(self.entity.apellidop))
        return items

    def all_data(self, idu):
        privilegios = UsuarioManager(self.db).get_privileges(idu, '/postulante')
        datos = self.db.query(self.entity).filter(self.entity.enabled).order_by(asc(self.entity.apellidop)).all()
        list = []

        for item in datos:
            is_active = item.estado

            disable = 'disabled' if 'postulante_update' not in privilegios else ''
            estado = 'Activo' if is_active else 'Inactivo'
            check = 'checked' if is_active else ''
            delete = 'postulante_delete' in privilegios

            if item.fkcargo:

                cargo = item.cargo.nombre
                delete = ''
                disable = 'disabled'
            else:
                cargo = item.tipo




            diccionario = item.get_dict()
            diccionario['estado'] = estado
            diccionario['check'] = check
            diccionario['disable'] = disable
            diccionario['delete'] = delete
            diccionario['fullname'] = item.fullname
            diccionario['cargo'] = cargo
            diccionario['fechar'] = item.fechar.strftime('%d/%m/%Y')

            list.append(diccionario)

        return list

    def insert(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()
        objeto.tipo = "POSTULANTE"
        objeto.fechar = fecha

        a = super().insert(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Registró postulante.", fecha=fecha, tabla="almacen_postulante", identificador=a.id)
        super().insert(b)
        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modificó postulante.", fecha=fecha, tabla="almacen_postulante", identificador=a.id)
        super().insert(b)
        return a


    def state(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        mensaje = "Habilitó postulante" if estado else "Deshabilitó postulante"
        x.estado = estado

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion=mensaje, fecha=fecha, tabla="almacen_postulante", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

    def delete(self, id, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = False

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó postulante", fecha=fecha, tabla="almacen_postulante", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()



class PostulanteTipoManager(SuperManager):

    def __init__(self, db):
        super().__init__(PostulanteTipo, db)


    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled)
        return items



class PostulanteTallaManager(SuperManager):

    def __init__(self, db):
        super().__init__(PostulanteTalla, db)


    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled)
        return items

class PostulanteColorManager(SuperManager):
    def __init__(self, db):
        super().__init__(PostulanteColor, db)

    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled)
        return items