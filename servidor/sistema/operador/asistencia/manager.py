from servidor.common.managers import SuperManager
from servidor.sistema.usuarios.bitacora.manager import BitacoraManager
from servidor.sistema.usuarios.usuario.manager import UsuarioManager
from servidor.sistema.usuarios.bitacora.model import Bitacora
from servidor.sistema.operador.asistencia.model import Asistencia,TipoAusencia
from servidor.sistema.operador.cliente.model import Cliente
from servidor.sistema.recursos_humanos.personal.model import Personal

from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from dateutil.relativedelta import relativedelta

import pytz
from sqlalchemy.sql import func
from sqlalchemy import func, desc, asc,or_,and_
from openpyxl import Workbook
from openpyxl.styles import Border , Alignment, Font, PatternFill,Side
from openpyxl.worksheet.table import Table, TableStyleInfo



global fecha_zona
fecha_zona = datetime.now(pytz.timezone('America/La_Paz'))


class AsistenciaManager(SuperManager):

    def __init__(self, db):
        super().__init__(Asistencia, db)


    def rango_fechas(self,fechai, fechaf):
        rango = []
        dias_totales = (fechaf - fechai).days
        for days in range(dias_totales + 1):
            fecha = fechai + relativedelta(days=days)
            rango.append(fecha)
        return rango

    def asistencia_excel(self,diccionary):
        fecha = datetime.now()
        cname = "Asistencia" + fecha.strftime('%Y-%m-%d') + ".xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = 'Reporte de Asistencia'

        indice = 0

        verdeFill = PatternFill(start_color='C6E0B4',
                               end_color='C6E0B4',
                               fill_type='solid')

        rojoFill = PatternFill(start_color='FF6565',
                                end_color='FF6565',
                                fill_type='solid')

        amarilloFill = PatternFill(start_color='FFFF89',
                               end_color='FFFF89',
                               fill_type='solid')

        azulFill = PatternFill(start_color='82A1D8',
                                   end_color='82A1D8',
                                   fill_type='solid')

        rosadoFill = PatternFill(start_color='FFB3B3',
                                   end_color='FFB3B3',
                                   fill_type='solid')

        narajaFill = PatternFill(start_color='FFD13F',
                                   end_color='FFD13F',
                                   fill_type='solid')

        cafeFill = PatternFill(start_color='DE6614',
                                 end_color='DE6614',
                                 fill_type='solid')

        celesteFill = PatternFill(start_color='61D6FF',
                                 end_color='61D6FF',
                                 fill_type='solid')

        plomoFill = PatternFill(start_color='D9D9D9',
                                  end_color='D9D9D9',
                                  fill_type='solid')

        blueFill = PatternFill(start_color='83C8F5',
                               end_color='83C8F5',
                               fill_type='solid')

        greenFill = PatternFill(start_color='8EF19D',
                                end_color='8EF19D',
                                fill_type='solid')

        yellowFill = PatternFill(start_color='FAFF88',
                                 end_color='FAFF88',
                                 fill_type='solid')

        # --------------------------------------------------------------------
        indice = indice + 1
        ws['A' + str(indice)] = 'PERSONAL'
        indice = indice + 2

        ws['A' + str(indice)] = 'Nº'
        ws['B' + str(indice)] = 'CI'
        ws['C' + str(indice)] = 'FECHA INGRESO'
        ws['D' + str(indice)] = 'FECHA DE NACIMIENTO'
        ws['E' + str(indice)] = 'CARGO'
        ws['F' + str(indice)] = 'TELEFONO'
        ws['G' + str(indice)] = 'NOMBRE COMPLETO'

        ws['A' + str(indice)].font = Font(bold=True)
        ws['B' + str(indice)].font = Font(bold=True)
        ws['C' + str(indice)].font = Font(bold=True)
        ws['D' + str(indice)].font = Font(bold=True)
        ws['E' + str(indice)].font = Font(bold=True)
        ws['F' + str(indice)].font = Font(bold=True)
        ws['G' + str(indice)].font = Font(bold=True)

        ws.column_dimensions['A'].height = 1
        ws.column_dimensions['A'].width = 5

        ws.column_dimensions['B'].height = 1
        ws.column_dimensions['B'].width = 14

        ws.column_dimensions['C'].height = 1
        ws.column_dimensions['C'].width = 16

        ws.column_dimensions['D'].height = 1
        ws.column_dimensions['D'].width = 16

        ws.column_dimensions['E'].height = 1
        ws.column_dimensions['E'].width = 25

        ws.column_dimensions['F'].height = 1
        ws.column_dimensions['F'].width = 12

        ws.column_dimensions['G'].height = 1
        ws.column_dimensions['G'].width = 40

        array_indices = ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA',
                         'AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP',
                         'AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

        fechainicio = datetime.strptime(diccionary['fecha-inicio-reporte'], '%d/%m/%Y')
        fechafin = datetime.strptime(diccionary['fecha-fin-reporte'], '%d/%m/%Y')
        lista_fechas = AsistenciaManager(self.db).rango_fechas(fechainicio, fechafin)
        cont_fecha = 0
        for fechadia in lista_fechas:
            ws.column_dimensions[array_indices[cont_fecha]].height = 1
            ws.column_dimensions[array_indices[cont_fecha]].width = 11

            ws[array_indices[cont_fecha] + str(indice)] = fechadia.strftime('%d/%m/%Y')
            ws[array_indices[cont_fecha] + str(indice)].font = Font(bold=True)
            cont_fecha = cont_fecha+ 1

        tipo_asistencia = self.db.query(TipoAusencia).filter(TipoAusencia.estado).filter(TipoAusencia.enabled).order_by(TipoAusencia.id.asc()).all()


        list_dict_tipo_asistencia = []

        for tipo_asist in tipo_asistencia:
            list_dict_tipo_asistencia.append(dict(tipo_asistencia=tipo_asist.codigo,cantidad=0))

            ws.column_dimensions[array_indices[cont_fecha]].height = 1
            ws.column_dimensions[array_indices[cont_fecha]].width = 7

            ws[array_indices[cont_fecha] + str(indice)] = tipo_asist.nombre
            ws[array_indices[cont_fecha] + str(indice)].font = Font(bold=True)
            ws[array_indices[cont_fecha] + str(indice)].alignment = Alignment(textRotation=90)

            if tipo_asist.codigo == "PR" :
                ws[array_indices[cont_fecha] + str(indice)].fill = verdeFill
            elif tipo_asist.codigo == "F":
                ws[array_indices[cont_fecha] + str(indice)].fill = rojoFill
            elif tipo_asist.codigo == "L":
                ws[array_indices[cont_fecha] + str(indice)].fill = amarilloFill
            elif tipo_asist.codigo == "X":
                ws[array_indices[cont_fecha] + str(indice)].fill = narajaFill
            elif tipo_asist.codigo == "BJM":
                ws[array_indices[cont_fecha] + str(indice)].fill = cafeFill
            elif tipo_asist.codigo == "V":
                ws[array_indices[cont_fecha] + str(indice)].fill = rosadoFill
            elif tipo_asist.codigo == "R":
                ws[array_indices[cont_fecha] + str(indice)].fill = azulFill
            elif tipo_asist.codigo == "PSG":
                ws[array_indices[cont_fecha] + str(indice)].fill = celesteFill

            cont_fecha = cont_fecha + 1

        if len(diccionary['personal']) == 0:
            personal = self.db.query(Personal).filter(Personal.estado).filter(Personal.enabled).order_by(
                Personal.apellidop.asc()).all()
        else:
            personal = self.db.query(Personal).filter(Personal.estado).filter(Personal.enabled).order_by(
                Personal.apellidop.asc()).filter(
                Personal.id.in_(diccionary['personal'])).all()

        for per in personal:
            indice = indice + 1
            ws['A' + str(indice)] = str(per.id)
            ws['B' + str(indice)] = str(per.ci)
            ws['C' + str(indice)] = per.fechar.strftime('%d/%m/%Y')
            ws['D' + str(indice)] = per.fechanacimiento.strftime('%d/%m/%Y')
            ws['E' + str(indice)] = str(per.cargo.nombre)
            ws['F' + str(indice)] = str(per.telefono)
            ws['G' + str(indice)] = str(per.apellidop) + ' ' + str(per.apellidom) + ' ' +str(per.nombre)

            cont_fecha = 0

            for x in range(len(list_dict_tipo_asistencia)):

                list_dict_tipo_asistencia[x]['cantidad'] = ''

            for fechadia in lista_fechas:

                asistencia = self.db.query(Asistencia).filter(Asistencia.fkpersonal == per.id) \
                    .filter(func.date(Asistencia.fechar) == func.date(fechadia)).first()

                if asistencia:
                    if asistencia.tipoausencia.nombre == "PRESENTE":

                        list_dict_tipo_asistencia[0]['cantidad'] = (0 if list_dict_tipo_asistencia[0]['cantidad'] == '' else int(list_dict_tipo_asistencia[0]['cantidad'])) + 1

                        if asistencia.fkturno == 1:

                            if asistencia.fkreemplazo:
                                ws[array_indices[cont_fecha] + str(indice)] = asistencia.reemplazo.codigo
                            else:
                                ws[array_indices[cont_fecha] + str(indice)] = asistencia.cliente.codigo

                        else:
                            if asistencia.fkreemplazo:
                                ws[array_indices[cont_fecha] + str(indice)] = asistencia.reemplazo.codigo+"-N"
                            else:
                                ws[array_indices[cont_fecha] + str(indice)] = asistencia.cliente.codigo+"-N"

                            ws[array_indices[cont_fecha] + str(indice)].fill = plomoFill

                    else:
                        ws[array_indices[cont_fecha] + str(indice)] = asistencia.tipoausencia.codigo
                        if asistencia.tipoausencia.codigo == "F":
                            ws[array_indices[cont_fecha] + str(indice)].fill = rojoFill
                            list_dict_tipo_asistencia[1]['cantidad'] = (0 if list_dict_tipo_asistencia[1]['cantidad'] == '' else int(list_dict_tipo_asistencia[1]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "L":
                            ws[array_indices[cont_fecha] + str(indice)].fill = amarilloFill
                            list_dict_tipo_asistencia[2]['cantidad'] = (0 if list_dict_tipo_asistencia[2]['cantidad'] == '' else int(list_dict_tipo_asistencia[2]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "X":
                            ws[array_indices[cont_fecha] + str(indice)].fill = narajaFill
                            list_dict_tipo_asistencia[3]['cantidad'] = (0 if list_dict_tipo_asistencia[3]['cantidad'] == '' else int(list_dict_tipo_asistencia[3]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "BJM":
                            ws[array_indices[cont_fecha] + str(indice)].fill = cafeFill
                            list_dict_tipo_asistencia[4]['cantidad'] = (0 if list_dict_tipo_asistencia[4]['cantidad'] == '' else int(list_dict_tipo_asistencia[4]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "V":
                            ws[array_indices[cont_fecha] + str(indice)].fill = rosadoFill
                            list_dict_tipo_asistencia[5]['cantidad'] = (0 if list_dict_tipo_asistencia[5]['cantidad'] == '' else int(list_dict_tipo_asistencia[5]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "R":
                            ws[array_indices[cont_fecha] + str(indice)].fill = azulFill
                            list_dict_tipo_asistencia[6]['cantidad'] = (0 if list_dict_tipo_asistencia[6]['cantidad'] == '' else int(list_dict_tipo_asistencia[6]['cantidad']))+ 1
                        elif asistencia.tipoausencia.codigo == "PSG":
                            ws[array_indices[cont_fecha] + str(indice)].fill = celesteFill
                            list_dict_tipo_asistencia[7]['cantidad'] = (0 if list_dict_tipo_asistencia[7]['cantidad'] == '' else int(list_dict_tipo_asistencia[7]['cantidad']))+ 1

                    ws[array_indices[cont_fecha] + str(indice)].font = Font(bold=True)

                else:
                    ws[array_indices[cont_fecha] + str(indice)] = ""


                cont_fecha = cont_fecha + 1

            for tip_asistencia in list_dict_tipo_asistencia:
                ws[array_indices[cont_fecha] + str(indice)] = tip_asistencia['cantidad']
                cont_fecha = cont_fecha + 1


        len_fechas = len(lista_fechas)
        #
        # for col in ws.columns:
        #
        #     column = col[0].column_letter
        #     ws.column_dimensions[column].height = 1
        #
        #     if column == "A" :
        #         ws.column_dimensions[column].width = 5
        #     elif column == "B":
        #         ws.column_dimensions[column].width = 14
        #     elif column == "C":
        #         ws.column_dimensions[column].width = 16
        #     elif column == "D":
        #         ws.column_dimensions[column].width = 16
        #     elif column == "E":
        #         ws.column_dimensions[column].width = 22
        #     elif column == "F":
        #         ws.column_dimensions[column].width = 12
        #     elif column == "G":
        #         ws.column_dimensions[column].width = 40
        #     else:
        #         ws.column_dimensions[column].width = 11


        wb.save("servidor/common/resources/downloads/" + cname)

        return cname


    def listar_habilitados(self):
        return self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).all()

    def listar_todo(self):
        return self.db.query(self.entity).filter(self.entity.enabled).all()

    def list_all(self):
        return dict(objects=self.db.query(self.entity).filter(self.entity.enabled))

    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled)
        return items

    # def all_data(self, idu):
    #     privilegios = UsuarioManager(self.db).get_privileges(idu, '/asistencia')
    #     datos = self.db.query(self.entity).filter(self.entity.enabled).all()
    #     list = []
    #
    #     for item in datos:
    #         is_active = item.estado
    #         # disable = 'disabled' if 'asistencia_update' not in privilegios else ''
    #         disable = ''
    #         estado = 'Activo' if is_active else 'Inactivo'
    #         check = 'checked' if is_active else ''
    #         delete = 'asistencia_delete' in privilegios
    #
    #
    #         diccionario = item.get_dict()
    #         diccionario['estado'] = estado
    #         diccionario['check'] = check
    #         diccionario['disable'] = disable
    #         diccionario['delete'] = delete
    #         diccionario['fechar'] = item.fechar.strftime('%d/%m/%Y %H:%M')
    #         diccionario['nombre'] = item.personal.fullname
    #         diccionario['codigo'] = item.cliente.dia if item.fkcliente else ''
    #         diccionario['cliente'] = item.cliente.nombre if item.fkcliente else ''
    #         diccionario['ausencia'] = item.tipoausencia.nombre if item.fktipoausencia else ''
    #
    #         list.append(diccionario)
    #
    #     return list

    def all_data(self, idu):
        privilegios = UsuarioManager(self.db).get_privileges(idu, '/asistencia')
        datos = self.db.query(Cliente).filter(Cliente.enabled).all()
        list = []

        for item in datos:
            is_active = item.estado
            disable = ''
            estado = 'Activo' if is_active else 'Inactivo'
            check = 'checked' if is_active else ''
            delete = 'asistencia_delete' in privilegios

            diccionario = item.get_dict()
            diccionario['estado'] = estado
            diccionario['check'] = check
            diccionario['disable'] = disable
            diccionario['delete'] = delete

            list.append(diccionario)

        return list

    def insert(self, dict):
        fzona = datetime.now(pytz.timezone('America/La_Paz'))
        hora_str = fzona.strftime('%H:%M')
        dict['fechar']  =dict['fechar'] + " "+hora_str

        dict['fechar'] = datetime.strptime(dict['fechar'], '%d/%m/%Y %H:%M')


        asistencia_personal = self.db.query(Asistencia).filter(Asistencia.fkpersonal == dict['fkpersonal']) \
            .filter(Asistencia.fkturno == dict['fkturno']) \
            .filter(or_(Asistencia.fkcliente == dict['fkcliente'],Asistencia.fkreemplazo == dict['fkcliente'])) \
            .filter(func.date(Asistencia.fechar) == func.date(dict['fechar'])).first()

        if dict['fktipoausencia'] == "":
            dict['fktipoausencia'] = None

        if dict['fkcliente'] == "":
            dict['fkcliente'] = None


        if asistencia_personal:
            objeto = asistencia_personal
            objeto.fktipoausencia = dict['fktipoausencia']

            if len(dict) != 7:

                if int(dict['fkreemplazo']) == objeto.fkcliente:
                    objeto.fkreemplazo = None
                else:
                    objeto.fkreemplazo = dict['fkreemplazo']
                    objeto.fktipoausencia = 1

            else:

                if dict['fktipoausencia'] != "1":
                    objeto.fkreemplazo = None

            a = super().update(objeto)
            b = Bitacora(fkusuario=dict['user'], ip=dict['ip'], accion="Modifico asistencia.", fecha=dict['fechar'],
                         tabla="almacen_asistencia", identificador=a.id)
            super().insert(b)
        else:
            objeto = self.entity(**dict)
            a = super().insert(objeto)
            b = Bitacora(fkusuario=dict['user'], ip=dict['ip'], accion="Registró asistencia.", fecha=dict['fechar'],
                         tabla="almacen_asistencia", identificador=a.id)
            super().insert(b)

        return a

    def update(self, objeto):
        fecha = BitacoraManager(self.db).fecha_actual()

        a = super().update(objeto)
        b = Bitacora(fkusuario=objeto.user, ip=objeto.ip, accion="Modificó asistencia.", fecha=fecha, tabla="almacen_asistencia", identificador=a.id)
        super().insert(b)
        return a


    def state(self, id, estado, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        mensaje = "Habilitó asistencia" if estado else "Deshabilitó asistencia"
        x.estado = estado

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion=mensaje, fecha=fecha, tabla="almacen_asistencia", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()
        return x

    def delete(self, id, user, ip):
        x = self.db.query(self.entity).filter(self.entity.id == id).one()
        x.enabled = False

        fecha = BitacoraManager(self.db).fecha_actual()
        b = Bitacora(fkusuario=user, ip=ip, accion="Eliminó asistencia", fecha=fecha, tabla="almacen_asistencia", identificador=id)
        super().insert(b)
        self.db.merge(x)
        self.db.commit()



class TipoAusenciaManager(SuperManager):

    def __init__(self, db):
        super().__init__(TipoAusencia, db)

    def listar_todo(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).order_by(self.entity.id.asc()).all()
        return items

    def get_all(self):
        items = self.db.query(self.entity).filter(self.entity.estado).filter(self.entity.enabled).order_by(self.entity.id.asc())
        return items



